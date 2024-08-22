import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox, Toplevel, Label, Entry, Button
from PIL import Image
import pytesseract
import pandas as pd
import cv2
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

# Function to open image files
def open_images():
    global extracted_texts
    file_paths = filedialog.askopenfilenames(filetypes=[("Image Files", "*.jpg;*.jpeg;*.png;*.bmp")])
    if file_paths:
        extracted_texts = []
        for file_path in file_paths:
            img = cv2.imread(file_path)
            extracted_text = pytesseract.image_to_string(img)
            extracted_texts.append(extracted_text)
        open_describe_window()

# Function to open the "Describe" window
def open_describe_window():
    describe_window = Toplevel(root)
    describe_window.title("Describe")
    describe_window.geometry("400x200")

    Label(describe_window, text="Enter the headers (comma separated):").pack(pady=10)
    headers_entry = Entry(describe_window, width=50)
    headers_entry.pack(pady=10)

    button_frame = tk.Frame(describe_window)
    button_frame.pack(pady=10)

    ok_btn = Button(button_frame, text="Ok", command=lambda: save_to_excel(headers_entry.get(), describe_window),
                    bg="#4CAF50", fg="white", width=10)
    ok_btn.pack(side="left", padx=5)

    cancel_btn = Button(button_frame, text="Cancel", command=describe_window.destroy,
                        bg="#f44336", fg="white", width=10)
    cancel_btn.pack(side="left", padx=5)

# Function to save the extracted data into an Excel file
def save_to_excel(headers, window):
    if headers:
        headers = [header.strip() for header in headers.split(",")]
        
        # Create a new Excel workbook and add a worksheet
        workbook = Workbook()
        sheet = workbook.active

        # Write headers to the first row
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f'{col_letter}1']
            cell.value = header
            cell.font = Font(bold=True)

        # Define border style
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        # Process each extracted text and write to Excel
        for row_num, extracted_text in enumerate(extracted_texts, start=2):
            lines = extracted_text.split('\n')
            data = []
            for header in headers:
                found = False
                for line in lines:
                    if header.lower() in line.lower():
                        data.append(line.split(':')[1].strip())
                        found = True
                        break
                if not found:
                    data.append("")
            
            # Write data to the row
            for col_num, value in enumerate(data, 1):
                col_letter = get_column_letter(col_num)
                cell = sheet[f'{col_letter}{row_num}']
                cell.value = value
                cell.border = thin_border

        # Auto-adjust column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Prompt the user to save the file
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            workbook.save(file_path)
            messagebox.showinfo("Success", "Data saved successfully!")
        window.destroy()

# Setting up the Tkinter window
root = tk.Tk()
root.title("Data Extraction Tool")
root.geometry("400x100")

# Buttons
btn_open = tk.Button(root, text="Upload Image", command=open_images, bg="#008CBA", fg="white", width=15)
btn_open.pack(padx=20, pady=20)

# Start the GUI loop
root.mainloop()
